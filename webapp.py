"""Flask application for the Digitization Process Management System."""

import csv
import os
from io import StringIO, BytesIO
from datetime import datetime

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, make_response
import sqlite3
from functools import wraps
from database import (
    run_startup,
    add_user,
    authenticate_user,
    list_users,
    get_user_by_email,
    update_user_account,
    update_user_status,
    delete_user_account,
    admin_reset_user_password,
    update_user_password,
    add_document,
    list_documents,
    get_document,
    add_process_tracking,
    update_document_details,
    delete_document,
    list_document_updates,
    list_status_counts,
    list_collection_options,
    replace_collection_options,
)

# Use pages/ as Jinja template folder to match current project structure.
app = Flask(__name__, template_folder="pages")
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True
app.secret_key = os.getenv("SECRET_KEY", "dev-only-secret-change-me")

PROCESS_STATUSES = [
    "คัดเลือกเอกสาร",
    "สแกนเอกสาร",
    "ตรวจไฟล์ภาพสแกน",
    "ตกแต่งภาพสแกน",
    "สร้างและฝัง Metadata ใน PDF",
    "จัดเก็บ/สำรองไฟล์",
    "นำไฟล์เอกสารเข้าระบบคลังสารสนเทศดิจิทัล Metadata ใน PDF",
]

# Initialize database on startup.
# On Render, set DB_PATH to a persistent disk path such as /var/data/digitization.db.
DB_PATH = os.getenv("DB_PATH", "data/digitization.db")
conn = run_startup(DB_PATH)

ACCOUNT_STATUS_LABELS = {
    "Active": "กำลังใช้งาน (Active)",
    "Inactive": "ถูกระงับการใช้งาน (Inactive)",
}


def _parse_datetime(value):
    """Parse ISO-like datetime strings from SQLite into datetime objects."""
    if not value:
        return None

    text = str(value).strip()
    if not text or text == "-":
        return None

    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        pass

    prepared_text = text
    if prepared_text.endswith("น."):
        prepared_text = prepared_text[:-2].strip()

    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H.%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(prepared_text, pattern)
        except ValueError:
            continue
    return None


def format_display_date(value):
    """Render date as DD/MM/YYYY."""
    parsed = _parse_datetime(value)
    if parsed:
        return parsed.strftime("%d/%m/%Y")
    return "-" if not value else str(value)


def format_display_time(value):
    """Render time as HH.MM น."""
    parsed = _parse_datetime(value)
    if parsed:
        return parsed.strftime("%H.%M น.")
    return "-"


def format_report_datetime(value):
    """Render datetime as YYYY-MM-DD HH:MM for reports."""
    parsed = _parse_datetime(value)
    if parsed:
        return parsed.strftime("%Y-%m-%d %H:%M")
    return "-"


app.jinja_env.filters["display_date"] = format_display_date
app.jinja_env.filters["display_time"] = format_display_time
app.jinja_env.filters["report_datetime"] = format_report_datetime


def _sort_documents(docs, sort_order):
    """Apply shared sorting rules for document collections."""
    order = (sort_order or "new").strip().lower()
    if order == "old":
        return sorted(
            docs,
            key=lambda d: ((d.get("last_completed_at") or d.get("created_at") or ""), (d.get("file_name") or "").casefold()),
        )
    if order == "title":
        return sorted(
            docs,
            key=lambda d: ((d.get("title") or "").casefold(), (d.get("file_name") or "").casefold()),
        )
    return sorted(
        docs,
        key=lambda d: ((d.get("last_completed_at") or d.get("created_at") or ""), (d.get("file_name") or "").casefold()),
        reverse=True,
    )


def _filter_documents_by_status(docs, status_filter, current_user_email, current_user_name):
    """Apply shared status filtering rules for document collections."""
    selected = (status_filter or "").strip()
    if not selected:
        return docs

    if selected == "my_job":
        normalized_email = (current_user_email or "").strip().lower()
        normalized_name = (current_user_name or "").strip()
        filtered_docs = []
        for doc in docs:
            owner_name = (doc.get("user_name") or "").strip()
            latest_user_email = (doc.get("latest_user_email") or "").strip().lower()
            latest_user_name = (doc.get("latest_user_name") or "").strip()
            if (
                (normalized_email and latest_user_email == normalized_email)
                or (normalized_name and owner_name == normalized_name)
                or (normalized_name and latest_user_name == normalized_name)
            ):
                filtered_docs.append(doc)
        return filtered_docs

    return [d for d in docs if (d.get("current_status") or "") == selected]


def login_required(f):
    """Decorator to check if user is logged in."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Decorator to allow access for admin users only."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        if session.get("user_role") != "Admin":
            return "Forbidden", 403
        return f(*args, **kwargs)
    return decorated_function


# ==================== Authentication Routes ====================

@app.route("/", methods=["GET"])
def index():
    """Home page - redirect based on login status."""
    if "user_email" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/health", methods=["GET"])
def health():
    """Basic application health check."""
    return jsonify({"status": "ok"}), 200


@app.route("/login", methods=["GET", "POST"])
def login():
    """User login page."""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = authenticate_user(conn, email, password)
        if user:
            session["user_email"] = user["email"]
            session["user_name"] = user["user_name"]
            session["user_role"] = user["role"]
            # Keep current login password only for read-only reveal on settings page.
            session["current_password_plain"] = password
            return redirect(url_for("dashboard"))
        else:
            return render_template("login.html", error="อีเมลหรือรหัสผ่านไม่ถูกต้อง")

    return render_template("login.html")


@app.route("/logout", methods=["GET", "POST"])
def logout():
    """User logout."""
    session.clear()
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    """Public registration is disabled. Accounts are created by admin only."""
    return "Registration is disabled. Please contact admin.", 403


# ==================== Dashboard Routes ====================

@app.route("/dashboard", methods=["GET"])
@login_required
def dashboard():
    """Dashboard - show overall progress timeline and all documents with filtering."""
    # Get all documents
    all_docs = list_documents(conn)
    
    # Build status counts map
    status_counts_map = {}
    for doc in all_docs:
        status = doc.get("current_status") or "คัดเลือกเอกสาร"
        status_counts_map[status] = status_counts_map.get(status, 0) + 1
    
    # Build timeline data for each process status
    status_timeline = []
    for idx, status_name in enumerate(PROCESS_STATUSES):
        count = status_counts_map.get(status_name, 0)
        status_timeline.append({
            "status": status_name,
            "count": count,
            "is_top": idx % 2 == 0
        })
    
    # Apply search/filter/sort
    query = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "").strip()
    sort_order = request.args.get("sort", "new")
    
    docs = all_docs
    
    if query:
        docs = [
            d for d in docs
            if query in (d.get("title") or "").lower() 
            or query in (d.get("file_name") or "").lower()
            or query in (d.get("bib") or "").lower()
            or query in (d.get("call_number") or "").lower()
        ]
    
    docs = _filter_documents_by_status(
        docs,
        status_filter,
        session.get("user_email"),
        session.get("user_name"),
    )
    
    docs = _sort_documents(docs, sort_order)
    
    return render_template(
        "dashboard.html",
        status_timeline=status_timeline,
        total_documents=len(all_docs),
        documents=docs,
        process_statuses=PROCESS_STATUSES,
        query=query,
        status_filter=status_filter,
        sort_order=sort_order,
        user_name=session.get("user_name"),
    )


# ==================== Document Routes ====================

@app.route("/documents", methods=["GET"])
@login_required
def documents_list():
    """Display all documents."""
    docs = list_documents(conn)

    # Query controls for page-level search/filter/sort.
    query = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "").strip()
    sort_order = request.args.get("sort", "new")

    if query:
        docs = [
            d for d in docs
            if query in (d.get("title") or "").lower() or query in (d.get("file_name") or "").lower()
        ]

    docs = _filter_documents_by_status(
        docs,
        status_filter,
        session.get("user_email"),
        session.get("user_name"),
    )

    docs = _sort_documents(docs, sort_order)

    status_options = sorted({(d.get("current_status") or "") for d in list_documents(conn) if d.get("current_status")})

    return render_template(
        "documents_list.html",
        documents=docs,
        q=query,
        status_filter=status_filter,
        sort_order=sort_order,
        status_options=status_options,
        message=request.args.get("message", "").strip().lower(),
    )




def _build_report_data():
    """Build report rows with latest status/update metadata for each document."""
    docs = list_documents(conn)
    report_data = []
    for d in docs:
        latest_status = d.get("current_status", "-")
        latest_updates = list_document_updates(conn, d.get("file_name", ""))
        latest_user = "-"
        latest_date = "-"

        if latest_updates:
            latest_update = latest_updates[-1]
            latest_user = latest_update.get("user_name", "-")
            latest_date = latest_update.get("completed_at", "") or "-"

        report_data.append({
            "file_name": d.get("file_name", "-"),
            "bib": d.get("bib", "-"),
            "call_number": d.get("call_number", "-"),
            "collection": d.get("collection", "-"),
            "title": d.get("title", "-"),
            "publish_date": d.get("publish_date", "-"),
            "latest_status": latest_status,
            "latest_user": latest_user,
            "latest_date": latest_date,
        })
    return report_data


def _apply_report_filter(report_data, filter_type):
    """Apply sorting options for report preview/download."""
    if filter_type == "collection":
        return sorted(report_data, key=lambda x: (x["collection"] or "").lower())
    if filter_type == "status":
        status_order = {status: i for i, status in enumerate(PROCESS_STATUSES)}
        return sorted(report_data, key=lambda x: status_order.get(x["latest_status"], 999))
    if filter_type == "staff":
        return sorted(report_data, key=lambda x: (x["latest_user"] or "").lower())
    return report_data


def _report_filter_label(filter_type):
    """Return display label for selected report filter."""
    labels = {
        "all": "แสดงทั้งหมด",
        "collection": "เลือกตาม Collection (เรียงลำดับจากน้อยไปมาก)",
        "status": "เลือกตาม Job Status (เรียงลำดับจากน้อยไปมาก)",
        "staff": "เลือกตาม Staff (เรียงลำดับจากน้อยไปมาก)",
    }
    return labels.get(filter_type, labels["all"])


@app.route("/reports", methods=["GET"])
@login_required
def report_page():
    """Display the report page with preview and filtering options."""
    file_type = request.args.get("file_type", "csv").lower()
    filter_type = request.args.get("filter_type", "all").lower()

    if file_type not in {"csv", "xlsx", "pdf"}:
        file_type = "csv"
    if filter_type not in {"all", "collection", "status", "staff"}:
        filter_type = "all"

    report_data = _build_report_data()
    report_data = _apply_report_filter(report_data, filter_type)

    return render_template(
        "report.html",
        report_data=report_data,
        selected_file_type=file_type,
        selected_filter_type=filter_type,
        selected_filter_label=_report_filter_label(filter_type),
    )


@app.route("/reports/download", methods=["GET"])
@login_required
def download_report():
    """Download document report as CSV, XLSX, or PDF."""
    file_type = request.args.get("file_type", "csv").lower()
    filter_type = request.args.get("filter_type", "all").lower()

    if file_type not in {"csv", "xlsx", "pdf"}:
        file_type = "csv"
    if filter_type not in {"all", "collection", "status", "staff"}:
        filter_type = "all"

    report_data = _build_report_data()
    report_data = _apply_report_filter(report_data, filter_type)

    # Generate file based on file_type
    if file_type == "csv":
        return _generate_csv_report(report_data)
    elif file_type == "xlsx":
        return _generate_xlsx_report(report_data)
    elif file_type == "pdf":
        return _generate_pdf_report(report_data)
    else:
        return _generate_csv_report(report_data)


def _generate_csv_report(report_data):
    """Generate CSV report."""
    output = StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow([
        "No.", "File Name", "#BIB", "CallNo.", "Collection", "Title",
        "Published Year", "Job Status", "Staff", "Update Date&Time"
    ])
    
    # Write data rows
    for i, d in enumerate(report_data, start=1):
        writer.writerow([
            i,
            d.get("file_name", "-"),
            d.get("bib", "-"),
            d.get("call_number", "-"),
            d.get("collection", "-"),
            d.get("title", "-"),
            d.get("publish_date", "-"),
            d.get("latest_status", "-"),
            d.get("latest_user", "-"),
            d.get("latest_date", "-"),
        ])
    
    response = make_response(output.getvalue().encode("utf-8-sig"))
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=digitization_report.csv"
    return response


def _generate_xlsx_report(report_data):
    """Generate XLSX report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return _generate_csv_report(report_data)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Set column widths
    columns_width = [6, 15, 12, 12, 15, 18, 12, 18, 12, 20]
    for i, width in enumerate(columns_width, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Write headers
    headers = [
        "No.", "File Name", "#BIB", "CallNo.", "Collection", "Title",
        "Published Year", "Job Status", "Staff", "Update Date&Time"
    ]
    
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data rows
    for row_num, d in enumerate(report_data, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = d.get("file_name", "-")
        ws.cell(row=row_num, column=3).value = d.get("bib", "-")
        ws.cell(row=row_num, column=4).value = d.get("call_number", "-")
        ws.cell(row=row_num, column=5).value = d.get("collection", "-")
        ws.cell(row=row_num, column=6).value = d.get("title", "-")
        ws.cell(row=row_num, column=7).value = d.get("publish_date", "-")
        ws.cell(row=row_num, column=8).value = d.get("latest_status", "-")
        ws.cell(row=row_num, column=9).value = d.get("latest_user", "-")
        ws.cell(row=row_num, column=10).value = d.get("latest_date", "-")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=digitization_report.xlsx"
    return response


def _generate_pdf_report(report_data):
    """Generate PDF report."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageTemplate, Frame, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        # Fallback to CSV if reportlab not available
        return _generate_csv_report(report_data)
    
    # Create PDF in landscape mode for better table fit
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
    
    # Prepare table data
    table_data = [
        ["No.", "File Name", "#BIB", "CallNo.", "Collection", "Title",
         "Published Year", "Job Status", "Staff", "Update Date&Time"]
    ]
    
    for i, d in enumerate(report_data, start=1):
        table_data.append([
            str(i),
            d.get("file_name", "-"),
            d.get("bib", "-"),
            d.get("call_number", "-"),
            d.get("collection", "-"),
            d.get("title", "-"),
            str(d.get("publish_date", "-")),
            d.get("latest_status", "-"),
            d.get("latest_user", "-"),
            d.get("latest_date", "-"),
        ])
    
    # Create table
    table = Table(table_data, colWidths=[0.4*inch, 1*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch, 0.8*inch, 1*inch, 0.8*inch, 1.2*inch])
    
    # Apply table styling
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
    ])
    table.setStyle(style)
    
    # Build PDF
    doc.build([table])
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "attachment; filename=digitization_report.pdf"
    return response



@app.route("/documents/add", methods=["GET", "POST"])
@admin_required
def add_document_page():
    """Add a new document."""
    if request.method == "POST":
        file_name = request.form.get("file_name", "").strip()
        bib = request.form.get("bib", "").strip()
        call_number = request.form.get("call_number", "").strip()
        collection = request.form.get("collection", "").strip()
        title = request.form.get("title", "").strip()
        publish_date_raw = request.form.get("publish_date", "").strip()
        actor_name = (session.get("user_name") or "").strip()
        actor_email = (session.get("user_email") or "").strip().lower()

        try:
            publish_date = int(publish_date_raw) if publish_date_raw else None
            add_document(
                conn,
                file_name,
                actor_name,
                bib,
                call_number,
                collection,
                title,
                publish_date,
                "",
                actor_email,
            )
            return redirect(url_for("process_tracking_page", file_name=file_name, source="new", message="created"))
        except sqlite3.IntegrityError as e:
            error_msg = "ไม่สามารถเพิ่มเอกสารได้ เนื่องจากมีรหัสเอกสารนี้อยู่แล้ว"
            return render_template(
                "add_document.html",
                error=error_msg,
                actor_name=actor_name,
                collection_options=list_collection_options(conn),
            )
        except ValueError:
            return render_template(
                "add_document.html",
                error="ปีที่พิมพ์ต้องเป็นตัวเลข พ.ศ. เท่านั้น",
                actor_name=actor_name,
                collection_options=list_collection_options(conn),
            )

    return render_template(
        "add_document.html",
        actor_name=session.get("user_name"),
        collection_options=list_collection_options(conn),
        message=request.args.get("message", "").strip().lower(),
    )


@app.route("/documents/<file_name>", methods=["GET"])
@login_required
def view_document(file_name):
    """View document details."""
    doc = get_document(conn, file_name)
    if not doc:
        return "Document not found", 404

    updates = list_document_updates(conn, file_name)

    q = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "").strip()
    sort_order = request.args.get("sort", "new")

    if q:
        updates = [
            u for u in updates
            if q in (u.get("status") or "").lower() or q in (u.get("user_name") or "").lower()
        ]

    if status_filter == "my_job":
        current_user_email = session.get("user_email", "").strip().lower()
        filtered_updates = []
        for u in updates:
            user_email = (u.get("user_email") or "").strip().lower()
            if current_user_email and user_email == current_user_email:
                filtered_updates.append(u)
        updates = filtered_updates
    elif status_filter:
        updates = [u for u in updates if (u.get("status") or "") == status_filter]

    updates = sorted(
        updates,
        key=lambda u: ((u.get("completed_at") or u.get("created_at") or ""), u.get("transaction_id") or 0),
        reverse=(sort_order != "old"),
    )

    return render_template(
        "view_document.html",
        document=doc,
        updates=updates,
        process_statuses=PROCESS_STATUSES,
        q=q,
        status_filter=status_filter,
        sort_order=sort_order,
        message=request.args.get("message", "").strip().lower(),
        error=request.args.get("error", "").strip().lower(),
        can_delete_document=session.get("user_role") == "Admin",
    )


@app.route("/documents/<file_name>/delete", methods=["POST"])
@login_required
def delete_document_page(file_name):
    """Delete a document after confirming the current user's password."""
    if session.get("user_role") != "Admin":
        return "Forbidden", 403

    doc = get_document(conn, file_name)
    if not doc:
        return "Document not found", 404

    password = request.form.get("password", "")
    next_page = request.form.get("next_page", "view_document").strip().lower()

    current_email = (session.get("user_email") or "").strip().lower()
    if not current_email or not password:
        if next_page == "process_tracking":
            return redirect(url_for("process_tracking_page", file_name=file_name, source="update", error="delete_auth_failed"))
        return redirect(url_for("view_document", file_name=file_name, error="delete_auth_failed"))

    current_user = authenticate_user(conn, current_email, password)
    if not current_user:
        if next_page == "process_tracking":
            return redirect(url_for("process_tracking_page", file_name=file_name, source="update", error="delete_auth_failed"))
        return redirect(url_for("view_document", file_name=file_name, error="delete_auth_failed"))

    delete_document(conn, file_name)
    return redirect(url_for("documents_list", message="deleted_document"))


@app.route("/documents/<file_name>/process-tracking", methods=["GET", "POST"])
@login_required
def process_tracking_page(file_name):
    """Display and update process tracking page for new/existing document workflows."""
    doc = get_document(conn, file_name)
    if not doc:
        return "Document not found", 404

    source = request.args.get("source", "update").strip().lower()
    is_new_case = source == "new"

    if is_new_case and session.get("user_role") != "Admin":
        return "Forbidden", 403

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "edit_details" and session.get("user_role") == "Admin":
            publish_date_raw = request.form.get("publish_date", "").strip()
            publish_date = int(publish_date_raw) if (publish_date_raw and publish_date_raw.isdigit()) else None
            update_document_details(
                conn,
                file_name,
                request.form.get("user_name", "").strip() or doc.get("user_name", ""),
                request.form.get("bib", "").strip(),
                request.form.get("call_number", "").strip(),
                request.form.get("collection", "").strip(),
                request.form.get("title", "").strip(),
                publish_date,
                request.form.get("file_path", "").strip(),
            )
            return redirect(url_for("process_tracking_page", file_name=file_name, source="update", message="saved"))

        if action == "update_status":
            selected_status = request.form.get("status", "").strip()
            completed_at = request.form.get("completed_at", "").strip() or None
            note = request.form.get("note", "").strip()

            if selected_status in PROCESS_STATUSES:
                if not completed_at:
                    completed_at = datetime.utcnow().replace(microsecond=0).isoformat()

                add_process_tracking(
                    conn,
                    file_name,
                    selected_status,
                    completed_at,
                    session.get("user_name"),
                    note,
                    session.get("user_email"),
                )
                return redirect(url_for("process_tracking_page", file_name=file_name, source="update", message="saved"))

        return redirect(url_for("process_tracking_page", file_name=file_name, source="update", error="save_failed"))

    updates = list_document_updates(conn, file_name)
    updates_asc = sorted(updates, key=lambda u: u.get("transaction_id") or 0)

    latest_status = None
    latest_by_status = {}
    for item in updates_asc:
        status_name = (item.get("status") or "").strip()
        if status_name in PROCESS_STATUSES:
            latest_status = status_name
            latest_by_status[status_name] = item

    current_idx = PROCESS_STATUSES.index(latest_status) if latest_status in PROCESS_STATUSES else -1
    latest_update = updates_asc[-1] if updates_asc else None

    timeline = []
    for idx, status_name in enumerate(PROCESS_STATUSES):
        if is_new_case:
            state = "new"
        elif current_idx >= 0 and idx < current_idx:
            state = "past"
        elif current_idx >= 0 and idx == current_idx:
            state = "current"
        else:
            state = "future"

        timeline.append(
            {
                "index": idx + 1,
                "status": status_name,
                "state": state,
                "is_top": idx % 2 == 0,
                "status_update": latest_by_status.get(status_name),
            }
        )

    breadcrumb = "เอกสารทั้งหมด > เพิ่มเอกสารใหม่"
    if not is_new_case:
        breadcrumb = f"เอกสารทั้งหมด > {doc.get('title') or '-'} > อัปเดตรายละเอียดข้อมูล/สถานะ"

    return render_template(
        "process_tracking.html",
        document=doc,
        breadcrumb=breadcrumb,
        is_new_case=is_new_case,
        timeline=timeline,
        latest_status=latest_status or "-",
        latest_note=(latest_update.get("note") if latest_update else "") or "",
        process_statuses=PROCESS_STATUSES,
        collection_options=list_collection_options(conn),
        can_edit_detail=session.get("user_role") == "Admin",
        users=list_users(conn) if session.get("user_role") == "Admin" else [],
        message=request.args.get("message", "").strip().lower(),
        error=request.args.get("error", "").strip().lower(),
    )


# ==================== Management Routes ====================

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings_page():
    """Display settings page for all users."""
    current_email = session.get("user_email", "")
    current_role = (session.get("user_role") or "").strip()
    is_admin_user = current_role == "Admin"

    if request.method == "POST":
        return redirect(url_for("settings_page", error="password_readonly"))

    account = get_user_by_email(conn, current_email) if current_email else None
    if not account:
        account = {
            "user_name": session.get("user_name", "-"),
            "email": current_email or "-",
            "note": "",
            "created_at": "-",
        }

    created_at = account.get("created_at") or "-"
    if created_at and created_at != "-":
        created_at = str(created_at)

    return render_template(
        "setting.html",
        account=account,
        is_admin_user=is_admin_user,
        password_display=session.get("current_password_plain", "********"),
        created_at_display=created_at,
        message=request.args.get("message", "").strip().lower(),
        error=request.args.get("error", "").strip().lower(),
    )


@app.route("/system-management", methods=["GET", "POST"])
@admin_required
def system_management_page():
    """Display system management page for admins."""
    if request.method == "POST":
        raw_options = request.form.getlist("collection_options")
        replace_collection_options(conn, raw_options)
        return redirect(url_for("system_management_page", message="saved"))

    mode = request.args.get("mode", "view").strip().lower()
    is_edit_mode = mode == "edit"
    return render_template(
        "system_management.html",
        collection_options=list_collection_options(conn),
        is_edit_mode=is_edit_mode,
        message=request.args.get("message", "").strip().lower(),
    )


@app.route("/create-account", methods=["GET", "POST"])
@admin_required
def create_account_page():
    """Create a new user account page for admin users."""
    form_values = {
        "user_name": "",
        "email": "",
        "role": "Staff",
        "note": "",
    }
    error_message = ""

    if request.method == "POST":
        form_values["user_name"] = request.form.get("user_name", "").strip()
        form_values["email"] = request.form.get("email", "").strip().lower()
        form_values["role"] = request.form.get("role", "Staff").strip() or "Staff"
        form_values["note"] = request.form.get("note", "").strip()
        raw_password = request.form.get("password", "")

        try:
            add_user(
                conn,
                form_values["email"],
                form_values["user_name"],
                raw_password,
                form_values["role"],
                "Active",
                form_values["note"],
            )
            return redirect(url_for("user_management_page", message="created"))
        except (sqlite3.IntegrityError, ValueError):
            error_message = "ไม่สามารถเพิ่มบัญชีผู้ใช้ได้ กรุณาตรวจสอบชื่อบัญชี อีเมล และรหัสผ่าน"

    created_at_preview = datetime.utcnow().replace(microsecond=0).isoformat().replace("T", " ")
    return render_template(
        "create_account.html",
        created_at_preview=created_at_preview,
        error_message=error_message,
        form_values=form_values,
    )


@app.route("/user-management", methods=["GET", "POST"])
@admin_required
def user_management_page():
    """Display user management page for admins."""
    error_message = ""

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "create":
            try:
                add_user(
                    conn,
                    request.form.get("email", "").strip(),
                    request.form.get("user_name", "").strip(),
                    request.form.get("password", "").strip(),
                    request.form.get("role", "Staff").strip() or "Staff",
                    request.form.get("account_status", "Active").strip() or "Active",
                )
            except (sqlite3.IntegrityError, ValueError):
                error_message = "ไม่สามารถเพิ่มบัญชีผู้ใช้ได้ กรุณาตรวจสอบชื่อบัญชี อีเมล และรหัสผ่าน"
            else:
                return redirect(url_for("user_management_page", message="created"))

        if action == "edit":
            current_email = request.form.get("current_email", "").strip().lower()
            if current_email:
                try:
                    update_user_account(
                        conn,
                        current_email,
                        request.form.get("user_name", "").strip(),
                        request.form.get("role", "Staff").strip() or "Staff",
                        request.form.get("account_status", "Active").strip() or "Active",
                        request.form.get("password", ""),
                    )
                except (sqlite3.IntegrityError, ValueError):
                    error_message = "ไม่สามารถอัปเดตข้อมูลบัญชีได้ กรุณาตรวจสอบชื่อบัญชี อีเมล และรหัสผ่าน"
                else:
                    return redirect(url_for("user_management_page", message="updated"))

        if action == "toggle_status":
            email = request.form.get("email", "").strip().lower()
            target_status = request.form.get("target_status", "Inactive").strip()
            if email and email != (session.get("user_email") or "").strip().lower():
                update_user_status(conn, email, target_status)
            return redirect(url_for("user_management_page", message="status"))

        if action == "delete":
            email = request.form.get("email", "").strip().lower()
            if email and email != (session.get("user_email") or "").strip().lower():
                delete_user_account(conn, email)
            return redirect(url_for("user_management_page", message="deleted"))

        if action == "reset_password":
            email = request.form.get("email", "").strip().lower()
            if email and email != (session.get("user_email") or "").strip().lower():
                try:
                    admin_reset_user_password(conn, email)
                except ValueError:
                    error_message = "รหัสผ่านเริ่มต้นไม่ถูกต้อง"
                else:
                    return redirect(url_for("user_management_page", message="password_reset"))

    users = list_users(conn)
    query_text = request.args.get("q", "").strip()
    query = query_text.lower()
    status_filter = request.args.get("status", "all").strip().lower()
    sort_order = request.args.get("sort", "new").strip().lower()
    edit_email = request.args.get("edit", "").strip().lower()
    show_create_form = request.args.get("create", "0").strip() == "1"

    if query:
        users = [
            user for user in users
            if query in (user.get("user_name") or "").lower() or query in (user.get("email") or "").lower()
        ]

    if status_filter == "active":
        users = [user for user in users if (user.get("account_status") or "Active") == "Active"]
    elif status_filter == "inactive":
        users = [user for user in users if (user.get("account_status") or "Active") == "Inactive"]

    if sort_order == "old":
        users = sorted(users, key=lambda user: (user.get("created_at") or "", user.get("user_name") or ""))
    elif sort_order == "name":
        users = sorted(users, key=lambda user: (user.get("user_name") or ""))
    else:
        users = sorted(
            users,
            key=lambda user: (user.get("created_at") or "", user.get("user_name") or ""),
            reverse=True,
        )

    edit_user = next((user for user in list_users(conn) if (user.get("email") or "").lower() == edit_email), None)

    return render_template(
        "user_management.html",
        users=users,
        query=query_text,
        status_filter=status_filter,
        sort_order=sort_order,
        show_create_form=show_create_form,
        edit_user=edit_user,
        account_status_labels=ACCOUNT_STATUS_LABELS,
        message=request.args.get("message", "").strip().lower(),
        error_message=error_message,
    )


# ==================== API Routes ====================

@app.route("/api/users", methods=["GET"])
@login_required
def api_list_users():
    """API - List all users (JSON)."""
    users = list_users(conn)
    return jsonify([dict(u) for u in users])


@app.route("/api/documents", methods=["GET"])
@login_required
def api_list_documents():
    """API - List all documents (JSON)."""
    docs = list_documents(conn)
    return jsonify([dict(d) for d in docs])


@app.route("/api/documents/<file_name>", methods=["GET"])
@login_required
def api_get_document(file_name):
    """API - Get document by file_name (JSON)."""
    doc = get_document(conn, file_name)
    if not doc:
        return jsonify({"error": "Document not found"}), 404
    return jsonify(dict(doc))


@app.route("/api/documents/<file_name>/status", methods=["POST"])
@login_required
def api_update_status(file_name):
    """API - Update document status."""
    data = request.get_json()
    status = data.get("status", "").strip()
    completed_at = data.get("completed_at")

    try:
        add_process_tracking(
            conn,
            file_name,
            status,
            completed_at,
            session.get("user_name"),
            "",
            session.get("user_email"),
        )
        return jsonify({"success": True, "message": "Status updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/status-summary", methods=["GET"])
@login_required
def api_status_summary():
    """API - Get status summary (JSON)."""
    counts = list_status_counts(conn)
    return jsonify([dict(c) for c in counts])


# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors."""
    return render_template("404.html"), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors."""
    return render_template("500.html"), 500


# ==================== Main ====================

# if __name__ == "__main__":
#     app.run(debug=True, host="0.0.0.0", port=5000)